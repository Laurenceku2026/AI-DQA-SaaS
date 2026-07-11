"""Build fully English DFMEA template shells from Chinese templates."""
from __future__ import annotations

import os
import re
from typing import Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES = os.path.join(HERE, "templates")

CJK_RE = re.compile(r"[\u4e00-\u9fff]")

# Longest-first replacement pairs (shared + template-specific)
COMMON_REPLACEMENTS: List[Tuple[str, str]] = [
    ("措施优先级是以严重度、频度以及检测评级的综合为基础的，目的是为降低风险而对各项措施进行优先排序",
     "Action Priority is based on severity, occurrence, and detection ratings to prioritize risk reduction actions."),
    ("DFMEA 和PFMEA 的措施优先级(AP)", "DFMEA and PFMEA Action Priority (AP)"),
    ("设计失效模式及后果分析（DFMEA）  （标准表格）", "Design FMEA (DFMEA) — Standard Form"),
    ("过程失效模式及后果分析（PFMEA）  （标准表格）", "Process FMEA (PFMEA) — Standard Form"),
    ("系统分析    步骤1 - 策划与准备", "System Analysis    Step 1 — Planning & Preparation"),
    ("步骤2 - structural analysis （结构分析）", "Step 2 — Structural Analysis"),
    ("步骤3 - Functional analysis （功能分析）", "Step 3 — Functional Analysis"),
    ("步骤4 - Failure analysis （失效分析）", "Step 4 — Failure Analysis"),
    ("步骤5 - Riske analysis （风险分析）", "Step 5 — Risk Analysis"),
    ("步骤6 - Optimization （优化）", "Step 6 — Optimization"),
    ("1.对于上一较高级别要素和\\或最终用户的失效影响( FE )",
     "1. Failure Effect (FE) on higher-level element and/or end user"),
    ("1.对于上一高层级要素或最终用户的失效影响(FE)",
     "1. Failure Effect (FE) on higher-level element and/or end user"),
    ("对失效起因或失效模式的当前探测控制（DC)", "Current Detection Control (DC) for cause or mode"),
    ("下一低层级要素或特性的失效原因(FC)", "Failure Cause (FC) of next lower-level element/characteristic"),
    ("当前对失效起因的预防控制（PC)", "Current Prevention Control (PC) for failure cause"),
    ("2.过程步骤的功能和产品特性\n（量值为可选项）",
     "2. Process step function and product characteristics (value optional)"),
    ("2.过程步骤的功能和产品特性 / （量值为可选项）",
     "2. Process step function and product characteristics (value optional)"),
    ("3.过程工作要素的功能和过程特性", "3. Process work element function and process characteristics"),
    ("2.过程步骤的失效模式（FM）", "2. Process step failure mode (FM)"),
    ("3.工作要素的失效起因（FC)", "3. Work element failure cause (FC)"),
    ("关注要素的失效模式(FM)", "Focus element failure mode (FM)"),
    ("下一低层级的功能及要求或特性", "Next lower-level function/requirement or characteristic"),
    ("1.上一高层级功能及要求", "1. Higher-level function & requirements"),
    ("关注要素功能及要求", "Focus element function & requirements"),
    ("设计失效模式和后果分析", "Design Failure Mode and Effects Analysis"),
    ("潜在失效     起因/机理", "Potential failure cause / mechanism"),
    ("现行设计     控制预防", "Current design prevention control"),
    ("现行设计     控制探测", "Current design detection control"),
    ("潜在失效     模式", "Potential failure mode"),
    ("潜在失效    \u3000后果", "Potential failure effect"),
    ("潜在失效   \u3000后果", "Potential failure effect"),
    ("风\u3000险\u3000顺\u3000序\u3000数RPN", "RPN"),
    ("建  议\u3000\u3000 措  施", "Recommended actions"),
    ("采取的\u3000  措  施", "Actions taken"),
    ("责任及目标 完成日期", "Responsibility & target completion date"),
    ("措  施  结  果", "Action results"),
    ("严\u3000重\u3000度\u3000 S", "Severity (S)"),
    ("频\u3000度\u3000O", "Occurrence (O)"),
    ("探\u3000测\u3000度\u3000D", "Detection (D)"),
    ("第  1 页    共   页", "Page 1 of"),
    ("FMEA日期（编制）", "FMEA date (original)"),
    ("FMEA日期（编制）　　　", "FMEA date (original)"),
    ("（设计 FMEA）", "(Design FMEA)"),
    ("历史/变更授权\n(适用时}", "History / change authorization (if applicable)"),
    ("历史/变更授权 / (适用时}", "History / change authorization (if applicable)"),
    ("对产品或工 厂的影响度 非常高", "Very high impact on product or plant"),
    ("对产品或工 厂的影响度 中等", "Moderate impact on product or plant"),
    ("对产品或工 厂的影响度 高", "High impact on product or plant"),
    ("对产品或工 厂的影响度 低", "Low impact on product or plant"),
    ("没有可察觉到的影响", "No discernible effect"),
    ("非常高 - 非常低", "Very high - very low"),
    ("非常低 - 非常高", "Very low - very high"),
    ("低 - 非常低", "Low - very low"),
    ("对失效起因发生的预测", "Prediction of failure cause occurrence"),
    ("失效分析及降低风险", "Failure analysis & risk reduction"),
    ("3.下一低层级或特性类型", "3. Next lower level or characteristic type"),
    ("1.过程项目的功能", "1. Process item function"),
    ("措施优先级（AP）", "Action Priority (AP)"),
    ("DFMEA ID编号：", "DFMEA ID:"),
    ("PFMEA ID编号：", "PFMEA ID:"),
    ("DFMEA开始时间：", "DFMEA start date:"),
    ("DFMEA修订时间：", "DFMEA revision date:"),
    ("PFMEA开始时间：", "PFMEA start date:"),
    ("PFMEA修订时间：", "PFMEA revision date:"),
    ("设计责任人：", "Design responsible:"),
    ("项目名称：", "Project name:"),
    ("公司名称：", "Company name:"),
    ("工厂地址：", "Plant address:"),
    ("顾客名称：", "Customer name:"),
    ("年型/项目：", "Model year / project:"),
    ("跨职能小组：", "Cross-functional team:"),
    ("保密等级：", "Confidentiality level:"),
    ("过程职责：", "Process responsibility:"),
    ("目标\n完成日期", "Target\ncompletion date"),
    ("目标 / 完成日期", "Target / completion date"),
    ("责任人姓名", "Responsible person"),
    ("潜在失效起因/机理", "Potential failure cause / mechanism"),
    ("潜在失效模式", "Potential failure mode"),
    ("表单编号：", "Form no.:"),
    ("产品名称：", "Product name:"),
    ("FMEA编号", "FMEA no."),
    ("第  1 页", "Page"),
    ("共   页", "of"),
    ("型号/项目", "Model / project"),
    ("型号：", "Model:"),
    ("关键日期", "Key date"),
    ("编 制 人", "Prepared by"),
    ("核心小组", "Core team"),
    ("设计责任", "Design responsibility"),
    ("特殊特性", "Special characteristics"),
    ("现行过程", "Current process"),
    ("控制预防", "Prevention control"),
    ("控制探测", "Detection control"),
    ("严  重 度  S", "Severity (S)"),
    ("探 测 度  D", "Detection (D)"),
    ("频 数  O", "Occurrence (O)"),
    ("级 别", "Class"),
    ("1.上一高层级", "1. Higher level"),
    ("2.关注要素", "2. Focus element"),
    ("1.过程项", "1. Process item"),
    ("2.过程步骤", "2. Process step"),
    ("3.过程工作要素", "3. Process work element"),
    ("失效影响(FE)", "Failure effect (FE)"),
    ("失效原因(FC)", "Failure cause (FC)"),
    ("严重度(S)", "Severity (S)"),
    ("发生度(O)", "Occurrence (O)"),
    ("探测度(D)", "Detection (D)"),
    ("现行预防控制(PC)", "Current prevention control (PC)"),
    ("现行探测控制(DC)", "Current detection control (DC)"),
    ("措施优先级(AP)", "Action priority (AP)"),
    ("筛选器代码(可选)", "Filter code (optional)"),
    ("筛选符号", "Filter symbol"),
    ("目标完成日期", "Target completion date"),
    ("采取基于证据的措施", "Evidence-based actions taken"),
    ("完成日期", "Completion date"),
    ("预防措施", "Prevention actions"),
    ("探测措施", "Detection actions"),
    ("责任人", "Responsible"),
    ("状态", "Status"),
    ("措施证据", "Action evidence"),
    ("备注", "Remarks"),
    ("持续改善", "Continuous improvement"),
    ("系统分析", "System analysis"),
    ("问题#", "Issue #"),
    ("（可选）", "(optional)"),
    ("多功能小组", "Cross-functional team"),
    ("修订", "Revision"),
    ("影响", "Effect"),
    ("探测能力", "Detection capability"),
    ("非常高", "Very high"),
    ("子系统", "Subsystem"),
    ("系统", "System"),
    ("部件", "Part"),
    ("要求", "Requirement"),
    ("预防", "Prevention"),
    ("探测", "Detection"),
    ("高", "High"),
    ("中", "Medium"),
    ("低", "Low"),
    ("非常低", "Very low"),
]

SHEET_RENAMES = {
    "DFMEA标准表格": "DFMEA Standard Form",
    "PFMEA标准表格": "PFMEA Standard Form",
    "措施优先级准则": "Action Priority Criteria",
}


def _sorted_rules(extra: Iterable[Tuple[str, str]] = ()) -> List[Tuple[str, str]]:
    rules = list(COMMON_REPLACEMENTS) + list(extra)
    rules.sort(key=lambda item: len(item[0]), reverse=True)
    return rules


def _translate_text(text: str, rules: List[Tuple[str, str]]) -> str:
    out = str(text)
    for src, dst in rules:
        if src in out:
            out = out.replace(src, dst)
    return out


def _translate_sheet(ws: Worksheet, rules: List[Tuple[str, str]]) -> None:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and CJK_RE.search(value):
                cell.value = _translate_text(value, rules)


def _clear_rows_from(ws: Worksheet, start_row: int, max_col: int = 30) -> None:
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _build_template1() -> None:
    src = os.path.join(TEMPLATES, "模板1-DFMEA旧版.xlsx")
    dst = os.path.join(TEMPLATES, "Template-1-DFMEA-Legacy.xlsx")
    rules = _sorted_rules()
    wb = load_workbook(src)
    if "DFMEA" in wb.sheetnames:
        ws = wb["DFMEA"]
        _translate_sheet(ws, rules)
        _clear_rows_from(ws, 11)
    if "Sheet1" in wb.sheetnames:
        ws1 = wb["Sheet1"]
        _translate_sheet(ws1, rules)
        _clear_rows_from(ws1, 4)
    wb.save(dst)
    print("Wrote", dst)


def _build_template2() -> None:
    src = os.path.join(TEMPLATES, "模板2-DFMEA新版.xlsx")
    dst = os.path.join(TEMPLATES, "Template-2-DFMEA-New.xlsx")
    rules = _sorted_rules()
    wb = load_workbook(src)
    for old_name, new_name in SHEET_RENAMES.items():
        if old_name in wb.sheetnames:
            ws = wb[old_name]
            _translate_sheet(ws, rules)
            if old_name == "DFMEA标准表格":
                _clear_rows_from(ws, 12, max_col=28)
            ws.title = new_name
    wb.save(dst)
    print("Wrote", dst)


def _verify(path: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    count = 0
    samples: List[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and CJK_RE.search(value):
                    count += 1
                    if len(samples) < 8:
                        samples.append(f"{name}!{cell.coordinate}: {value[:60]}")
    wb.close()
    if samples:
        print("Remaining CJK in", os.path.basename(path))
        for line in samples:
            print(" ", line)
    return count


def main() -> None:
    _build_template1()
    _build_template2()
    for name in ("Template-1-DFMEA-Legacy.xlsx", "Template-2-DFMEA-New.xlsx"):
        remaining = _verify(os.path.join(TEMPLATES, name))
        print(name, "remaining CJK cells:", remaining)


if __name__ == "__main__":
    main()
