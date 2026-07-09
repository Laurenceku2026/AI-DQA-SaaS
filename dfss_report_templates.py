"""Fill client DFMEA Excel templates for AI-DQA exports."""
from __future__ import annotations

import os
import re
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

TEMPLATE_EXTENSIONS = (".xlsx", ".xls", ".docx")
DFMEA_SHEET_NAME = "DFMEA标准表格"
DFMEA_DATA_START_ROW = 12

# Header value cells in the client DFMEA template (1-based row/col).
DFMEA_HEADER_CELLS = {
    "project_name": (3, 17),  # Q3 项目名称
    "start_date": (4, 17),  # Q4 DFMEA开始时间
    "revision_date": (5, 17),  # Q5 DFMEA修订时间
    "customer_name": (5, 5),  # E5 顾客名称
    "model_year_project": (6, 5),  # E6 年型/项目
    "cross_function_team": (6, 9),  # I6 跨职能小组
}

# Risk row mapping for AI-DQA output -> DFMEA columns.
DFMEA_RISK_COLUMNS = {
    "higher_level": 4,
    "focus_element": 5,
    "lower_level": 6,
    "severity": 11,
    "failure_mode": 12,
    "cause": 13,
    "occurrence": 15,
    "detection": 17,
}


def list_report_templates(app_key: str = "AI-DQA") -> List[str]:
    """List bundled/customer report templates for user selection."""
    here = os.path.dirname(os.path.abspath(__file__))
    folders = [
        os.path.join(here, "templates"),
        os.path.join(os.environ.get("DFSS_TEMPLATE_DIR", ""), app_key),
        os.path.join(r"C:\Users\Laurence\Technical\Project\SaaS\DFSS Report Template", app_key),
    ]
    found: List[str] = []
    for folder in folders:
        if not folder or not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.startswith("~$"):
                continue
            if os.path.splitext(name)[1].lower() in TEMPLATE_EXTENSIONS:
                if name not in found:
                    found.append(name)
    return sorted(found)


def template_mime_type(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".xls":
        return "application/vnd.ms-excel"
    if ext == ".docx":
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export_report_template(
    template_filename: str,
    product_name: str,
    product_desc: str,
    report_content: str,
    analyst_name: str = "",
    analyst_title: str = "",
    lang: str = "zh",
) -> Tuple[BytesIO, str]:
    """Fill a selected report template. Fixed client templates use rule mapping (no LLM)."""
    lower_name = template_filename.lower()
    if lower_name.endswith((".xlsx", ".xls")) and ("fmea" in lower_name or "dfmea" in lower_name):
        data = fill_dfmea_template(
            product_name=product_name,
            product_desc=product_desc,
            report_content=report_content,
            analyst_name=analyst_name,
            analyst_title=analyst_title,
            lang=lang,
            template_filename=template_filename,
        )
        return data, template_mime_type(template_filename)
    raise ValueError(
        "当前模板暂未配置自动填表规则。"
        if lang == "zh"
        else "Automatic fill rules are not configured for this template yet."
    )


def resolve_template_path(filename: str, app_key: str = "AI-DQA") -> str:
    """Resolve bundled/local/customer template path."""
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(here, "templates", filename),
        os.path.join(os.environ.get("DFSS_TEMPLATE_DIR", ""), app_key, filename),
        os.path.join(
            r"C:\Users\Laurence\Technical\Project\SaaS\DFSS Report Template",
            app_key,
            filename,
        ),
    ]
    for path in candidates:
        if path and os.path.isfile(path):
            return path
    raise FileNotFoundError(f"DFMEA template not found: {filename}")


def _clean_cell_text(text: str) -> str:
    return re.sub(r"\*\*", "", text or "").strip()


def _normalize_header(header: str) -> str:
    h = _clean_cell_text(header).lower()
    mapping = {
        "模块": "module",
        "module": "module",
        "失效模式": "failure_mode",
        "failure mode": "failure_mode",
        "原因": "cause",
        "cause": "cause",
        "严重度": "severity",
        "severity": "severity",
        "发生度": "occurrence",
        "occurrence": "occurrence",
        "探测度": "detection",
        "detection": "detection",
        "rpn": "rpn",
    }
    for key, value in mapping.items():
        if key in h:
            return value
    return h


def parse_risks_from_markdown(report_content: str) -> List[Dict[str, str]]:
    """Parse Top-N risk table from AI-DQA markdown report."""
    risks: List[Dict[str, str]] = []
    header_map: Optional[Dict[str, int]] = None

    for line in report_content.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            if header_map and risks:
                break
            continue

        cells = [_clean_cell_text(c) for c in line.split("|")[1:-1]]
        if not cells:
            continue
        if all(re.fullmatch(r":?-{3,}:?", c.replace(" ", "")) for c in cells):
            continue

        normalized = [_normalize_header(c) for c in cells]
        if "module" in normalized and "failure_mode" in normalized:
            header_map = {name: idx for idx, name in enumerate(normalized)}
            continue

        if not header_map:
            continue

        def pick(field: str) -> str:
            idx = header_map.get(field)
            if idx is None or idx >= len(cells):
                return ""
            return cells[idx]

        module = pick("module")
        failure_mode = pick("failure_mode")
        if not module and not failure_mode:
            continue

        risks.append(
            {
                "module": module,
                "failure_mode": failure_mode,
                "cause": pick("cause"),
                "severity": pick("severity"),
                "occurrence": pick("occurrence"),
                "detection": pick("detection"),
                "rpn": pick("rpn"),
            }
        )

    return risks[:10]


def _set_cell(ws, row: int, col: int, value) -> None:
    if value is None:
        return
    text = str(value).strip()
    if not text:
        return
    ws.cell(row=row, column=col, value=text)


def _to_int(value: str):
    if value is None:
        return None
    m = re.search(r"\d+", str(value))
    return int(m.group()) if m else None


def fill_dfmea_template(
    product_name: str,
    product_desc: str,
    report_content: str,
    analyst_name: str = "",
    analyst_title: str = "",
    lang: str = "zh",
    template_filename: str = "新版FMEA表格.xlsx",
) -> BytesIO:
    """Fill the client DFMEA template and return an in-memory .xlsx file."""
    template_path = resolve_template_path(template_filename, "AI-DQA")
    wb = load_workbook(template_path)
    ws = wb[DFMEA_SHEET_NAME] if DFMEA_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    today = datetime.now().strftime("%Y-%m-%d")
    team = analyst_name or ("未填写" if lang == "zh" else "N/A")
    if analyst_title:
        team = f"{team} ({analyst_title})"

    header_values = {
        "project_name": product_name,
        "start_date": today,
        "revision_date": today,
        "customer_name": product_name,
        "model_year_project": product_desc[:80] if product_desc else product_name,
        "cross_function_team": team,
    }
    for key, (row, col) in DFMEA_HEADER_CELLS.items():
        _set_cell(ws, row, col, header_values.get(key))

    risks = parse_risks_from_markdown(report_content)
    if not risks:
        risks = [
            {
                "module": product_name,
                "failure_mode": "待补充失效模式" if lang == "zh" else "Pending failure mode",
                "cause": product_desc[:120] if product_desc else "",
                "severity": "",
                "occurrence": "",
                "detection": "",
                "rpn": "",
            }
        ]

    for idx, risk in enumerate(risks):
        row = DFMEA_DATA_START_ROW + idx
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["higher_level"], product_name)
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["focus_element"], risk.get("module"))
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["lower_level"], product_desc[:60] if product_desc else "-")
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["failure_mode"], risk.get("failure_mode"))
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["cause"], risk.get("cause"))
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["severity"], _to_int(risk.get("severity")))
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["occurrence"], _to_int(risk.get("occurrence")))
        _set_cell(ws, row, DFMEA_RISK_COLUMNS["detection"], _to_int(risk.get("detection")))

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
